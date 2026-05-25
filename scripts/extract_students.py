"""
Lee 'listado con numero de orden.xlsx' y produce data/students.csv
con columnas: matricula, grupo_codigo, grado, grupo, jornada, nombre.

Codigo del grupo: GGJJNN (grado, grupo, jornada)
  06 01 00 -> grado 6, grupo 1, jornada 00 (manana)
"""
import csv
import re
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent
XLSX = BASE / "listado con numero de orden.xlsx"
OUT_DIR = BASE / "data"
OUT_DIR.mkdir(exist_ok=True)
OUT_CSV = OUT_DIR / "students.csv"

JORNADA = {"00": "Manana", "01": "Tarde", "02": "Unica"}


def clean_name(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip().title()


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    rows = []
    for sn in wb.sheetnames:
        if not re.fullmatch(r"\d{6}", sn):
            continue
        grado = int(sn[0:2])
        grupo = int(sn[2:4])
        jor = sn[4:6]
        ws = wb[sn]
        for r in ws.iter_rows(min_row=11, values_only=True):
            num, matricula, name = r[1], r[2], r[4]
            if not isinstance(num, int) or matricula is None or not name:
                continue
            rows.append({
                "matricula": str(matricula).strip(),
                "grupo_codigo": sn,
                "grado": grado,
                "grupo": f"{grado}-{grupo}",
                "jornada": JORNADA.get(jor, jor),
                "nombre": clean_name(str(name)),
            })

    rows.sort(key=lambda x: (x["grupo_codigo"], x["nombre"]))
    with OUT_CSV.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["matricula", "grupo_codigo", "grado", "grupo", "jornada", "nombre"])
        w.writeheader()
        w.writerows(rows)

    print(f"Escrito: {OUT_CSV}  ({len(rows)} estudiantes)")
    by_grado = {}
    for x in rows:
        by_grado[x["grado"]] = by_grado.get(x["grado"], 0) + 1
    for g in sorted(by_grado):
        print(f"  grado {g}: {by_grado[g]}")


if __name__ == "__main__":
    main()
